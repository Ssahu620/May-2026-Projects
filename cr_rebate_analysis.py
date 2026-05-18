import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

INPUT_CR = "v2.Conversion Rate.xlsx"
INPUT_VC = "v2.2026 VC Master Data.xlsx"
OUTPUT   = "CR_Increase_No_Rebate_Analysis.xlsx"

MONTHS = ["Jan 2026", "Feb 2026", "Mar 2026", "Apr 2026", "May 2026"]

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY       = "1F3864"
BLUE       = "2E75B6"
LT_BLUE    = "D6E4F0"
WHITE      = "FFFFFF"
GRAY       = "F2F2F2"
DARK       = "1F1F1F"
GREEN_BG   = "E2EFDA"
GREEN_FG   = "375623"
GREEN_TAB  = "538135"
ORANGE_BG  = "FCE4D6"
ORANGE_FG  = "843C0C"
ORANGE_TAB = "C55A11"
BORDER_CLR = "BFBFBF"


# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def _font(bold=False, color=DARK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _border(color=BORDER_CLR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _auto_width(ws, min_w=10, max_w=52):
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        w = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[ltr].width = min(max(w + 2, min_w), max_w)


# ── Data functions ────────────────────────────────────────────────────────────
def load_data():
    cr_df = pd.read_excel(INPUT_CR, sheet_name="Sheet1")
    vc_df = pd.read_excel(INPUT_VC, sheet_name="Sheet1")
    return cr_df, vc_df


def find_cr_increases(cr_df):
    records = []
    for _, row in cr_df.iterrows():
        for i in range(1, len(MONTHS)):
            prev_m, curr_m = MONTHS[i - 1], MONTHS[i]
            pv, cv = row[prev_m], row[curr_m]
            if pd.notna(pv) and pd.notna(cv) and cv > pv:
                records.append({
                    "Item":                       row["Item"],
                    "ASIN":                       row["ASIN"],
                    "Brand":                      row["Brand"],
                    "Program-Category":           row["Program-Category"],
                    "Description | Color | Size": row["Description | Color | Size"],
                    "Month":                      curr_m,
                    "Prev Month":                 prev_m,
                    "Prev CR":                    round(pv, 4),
                    "Curr CR":                    round(cv, 4),
                    "CR Change":                  round(cv - pv, 4),
                })
    return pd.DataFrame(records)


def find_rebate_months(vc_df):
    mask = (vc_df["Deals Rebate"] > 0) | (vc_df["Coupons Rebate"] > 0)
    rm = (
        vc_df[mask][["Item", "MonthYear"]]
        .drop_duplicates()
        .rename(columns={"MonthYear": "Month"})
    )
    rm["has_rebate"] = True
    return rm


def split_by_rebate(cr_increases, rebate_months):
    cr = cr_increases.copy()
    rm = rebate_months.copy()
    cr["Item"] = cr["Item"].astype(float)
    rm["Item"] = rm["Item"].astype(float)
    merged = cr.merge(rm, on=["Item", "Month"], how="left")
    no_rebate   = merged[merged["has_rebate"].isna()].drop(columns="has_rebate")
    with_rebate = merged[merged["has_rebate"] == True].drop(columns="has_rebate")
    return no_rebate, with_rebate


def build_rebate_detail(vc_df):
    mask = (vc_df["Deals Rebate"] > 0) | (vc_df["Coupons Rebate"] > 0)
    return (
        vc_df[mask]
        .groupby(["Item", "MonthYear", "ASIN", "Description | Color | Size", "Brand"])
        .agg(
            Total_Deals_Rebate  =("Deals Rebate",   "sum"),
            Total_Coupons_Rebate=("Coupons Rebate", "sum"),
        )
        .reset_index()
        .sort_values("MonthYear")
    )


def build_brand_summary(no_rebate):
    return (
        no_rebate[~no_rebate["Brand"].astype(str).str.contains("TOTAL|SUB-TOTAL", na=False)]
        .groupby("Brand")
        .agg(
            Unique_Items         =("Item", "nunique"),
            CR_Increase_Instances=("Item", "count"),
        )
        .reset_index()
        .sort_values("CR_Increase_Instances", ascending=False)
    )


# ── Sheet builders ────────────────────────────────────────────────────────────
def _write_title_rows(ws, title, subtitle, n_cols):
    last_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = title
    c.fill      = _fill(NAVY)
    c.font      = _font(bold=True, color=WHITE, size=14)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = subtitle
    c.fill      = _fill(LT_BLUE)
    c.font      = _font(italic=True, color=DARK, size=10)
    c.alignment = _align("center")
    ws.row_dimensions[2].height = 16


def _write_header_row(ws, row, headers):
    ws.row_dimensions[row].height = 20
    for ci, label in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=label)
        c.fill      = _fill(BLUE)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.alignment = _align("center")
        c.border    = _border(WHITE)


def _write_data_row(ws, row, values, pct_cols, cur_cols, text_cols):
    bg = GRAY if (row % 2 == 0) else WHITE
    ws.row_dimensions[row].height = 15
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill      = _fill(bg)
        c.font      = _font(size=10)
        c.alignment = _align("left" if ci in text_cols else "center")
        c.border    = _border()
        if ci in pct_cols and isinstance(val, (int, float)):
            c.number_format = "0.00%"
        if ci in cur_cols and isinstance(val, (int, float)):
            c.number_format = '"$"#,##0.00'


def write_data_sheet(wb, df, sheet_name, title, subtitle, tab_color,
                     pct_col_names=None, cur_col_names=None, text_col_names=None):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color

    cols      = list(df.columns)
    n_cols    = len(cols)
    pct_cols  = {cols.index(c) + 1 for c in (pct_col_names  or []) if c in cols}
    cur_cols  = {cols.index(c) + 1 for c in (cur_col_names  or []) if c in cols}
    text_cols = {cols.index(c) + 1 for c in (text_col_names or []) if c in cols}

    _write_title_rows(ws, title, subtitle, n_cols)
    _write_header_row(ws, 3, cols)

    for ri, (_, row_data) in enumerate(df.iterrows(), 4):
        _write_data_row(ws, ri, list(row_data), pct_cols, cur_cols, text_cols)

    ws.freeze_panes = "A4"
    _auto_width(ws)


def write_executive_summary(wb, cr_increases, no_rebate, with_rebate, brand_summary):
    ws = wb.create_sheet("Executive Summary", 0)
    ws.sheet_properties.tabColor = NAVY

    today = date.today().strftime("%B %d, %Y")

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "Conversion Rate Organic Growth Analysis"
    c.fill      = _fill(NAVY)
    c.font      = _font(bold=True, color=WHITE, size=16)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value     = f"Items with CR Increase and No Deal or Coupon Rebate  |  Run Date: {today}"
    c.fill      = _fill(BLUE)
    c.font      = _font(italic=True, color=WHITE, size=11)
    c.alignment = _align("center")
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 12   # spacer

    # ── Key metric boxes ──────────────────────────────────────────────────────
    metrics = [
        ("Total CR\nIncrease\nInstances", len(cr_increases),          BLUE,      WHITE),
        ("CR Increased\nWith Rebate",     len(with_rebate),            ORANGE_BG, ORANGE_FG),
        ("CR Increased\nNo Rebate",       len(no_rebate),              GREEN_BG,  GREEN_FG),
        ("Unique Items\n(No Rebate)",     no_rebate["Item"].nunique(), GREEN_BG,  GREEN_FG),
    ]
    for i, (label, val, bg, fg) in enumerate(metrics):
        col = i + 1
        lc = ws.cell(row=4, column=col, value=label)
        lc.fill      = _fill(bg)
        lc.font      = _font(bold=True, color=fg, size=10)
        lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lc.border    = _border()
        ws.row_dimensions[4].height = 40

        vc_ = ws.cell(row=5, column=col, value=val)
        vc_.fill      = _fill(bg)
        vc_.font      = _font(bold=True, color=fg, size=26)
        vc_.alignment = _align("center")
        vc_.border    = _border()
        ws.row_dimensions[5].height = 44

    ws.row_dimensions[6].height = 12   # spacer

    # ── Brand summary table ───────────────────────────────────────────────────
    ws.merge_cells("A7:D7")
    c = ws["A7"]
    c.value     = "Organic CR Growth by Brand  (No Rebate Active)"
    c.fill      = _fill(NAVY)
    c.font      = _font(bold=True, color=WHITE, size=12)
    c.alignment = _align("center")
    ws.row_dimensions[7].height = 24

    hdrs = ["Brand", "Unique Items", "CR Increase Instances", "% of Items"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=8, column=ci, value=h)
        c.fill      = _fill(BLUE)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.alignment = _align("center")
        c.border    = _border(WHITE)
    ws.row_dimensions[8].height = 18

    total_items = brand_summary["Unique_Items"].sum()
    for ri, (_, row_) in enumerate(brand_summary.iterrows()):
        r   = 9 + ri
        bg  = GRAY if ri % 2 == 0 else WHITE
        pct = row_["Unique_Items"] / total_items if total_items else 0
        for ci, val in enumerate([row_["Brand"], int(row_["Unique_Items"]),
                                   int(row_["CR_Increase_Instances"]), pct], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(size=10)
            c.alignment = _align("left" if ci == 1 else "center")
            c.border    = _border()
            if ci == 4:
                c.number_format = "0%"
        ws.row_dimensions[r].height = 16

    tot_row = 9 + len(brand_summary)
    for ci, val in enumerate(["TOTAL", int(total_items),
                               int(brand_summary["CR_Increase_Instances"].sum()), 1.0], 1):
        c = ws.cell(row=tot_row, column=ci, value=val)
        c.fill      = _fill(NAVY)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.alignment = _align("left" if ci == 1 else "center")
        c.border    = _border(WHITE)
        if ci == 4:
            c.number_format = "0%"
    ws.row_dimensions[tot_row].height = 18

    # ── Methodology note ──────────────────────────────────────────────────────
    note_row = tot_row + 2
    ws.merge_cells(f"A{note_row}:F{note_row}")
    c = ws.cell(row=note_row, column=1,
                value=("Methodology: CR increase = current month conversion rate strictly greater "
                       "than prior month (Jan→Feb, Feb→Mar, Mar→Apr, Apr→May 2026). "
                       "Rebate check: Deals Rebate or Coupons Rebate > 0 in VC Master Data "
                       "for the same item and month."))
    c.fill      = _fill(LT_BLUE)
    c.font      = _font(italic=True, size=9, color="404040")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[note_row].height = 32

    # Column widths
    for col, w in zip("ABCDEF", [24, 14, 24, 14, 14, 14]):
        ws.column_dimensions[col].width = w


# ── Main output writer ────────────────────────────────────────────────────────
def save_output(no_rebate, with_rebate, rebate_detail, brand_summary, cr_increases):
    TEXT_COLS = ["Item", "ASIN", "Brand", "Program-Category",
                 "Description | Color | Size", "Month", "Prev Month"]
    try:
        wb = Workbook()
        wb.remove(wb.active)
    except Exception:
        wb = Workbook()

    write_executive_summary(wb, cr_increases, no_rebate, with_rebate, brand_summary)

    write_data_sheet(
        wb,
        no_rebate.sort_values(["Brand", "Item", "Month"]),
        sheet_name="CR Up - No Rebate",
        title="CR Increase — No Deal or Coupon Rebate",
        subtitle="Items where conversion rate improved month-over-month with no active rebate",
        tab_color=GREEN_TAB,
        pct_col_names=["Prev CR", "Curr CR", "CR Change"],
        text_col_names=TEXT_COLS,
    )

    write_data_sheet(
        wb,
        with_rebate.sort_values(["Brand", "Item", "Month"]),
        sheet_name="CR Up - Had Rebate",
        title="CR Increase — Rebate Was Active",
        subtitle="Items where conversion rate improved but a Deal or Coupon Rebate was active that month",
        tab_color=ORANGE_TAB,
        pct_col_names=["Prev CR", "Curr CR", "CR Change"],
        text_col_names=TEXT_COLS,
    )

    write_data_sheet(
        wb,
        rebate_detail,
        sheet_name="Rebate Detail",
        title="Deal & Coupon Rebate Detail",
        subtitle="All item-month combinations where Deals Rebate or Coupons Rebate > 0",
        tab_color=BLUE,
        cur_col_names=["Total_Deals_Rebate", "Total_Coupons_Rebate"],
        text_col_names=["Item", "MonthYear", "ASIN", "Description | Color | Size", "Brand"],
    )

    try:
        wb.save(OUTPUT)
    except PermissionError:
        raise PermissionError(
            f"Cannot write '{OUTPUT}' — please close the file in Excel and try again."
        )


# ── Console summary ───────────────────────────────────────────────────────────
def print_summary(cr_increases, no_rebate, with_rebate, brand_summary):
    print(f"\n{'='*55}")
    print("  Conversion Rate Increase + Rebate Analysis")
    print(f"{'='*55}")
    print(f"  Total CR increase instances (all):   {len(cr_increases):>5}")
    print(f"  CR increased WITH rebate:            {len(with_rebate):>5}")
    print(f"  CR increased WITHOUT rebate:         {len(no_rebate):>5}")
    print(f"  Unique items (no rebate):            {no_rebate['Item'].nunique():>5}")
    print(f"\n  Summary by Brand (no rebate):")
    print(f"  {'-'*45}")
    for _, r in brand_summary.iterrows():
        print(f"  {r['Brand']:<20} Items: {int(r['Unique_Items']):>3}   Instances: {int(r['CR_Increase_Instances']):>3}")
    print(f"\n  Output saved to: {OUTPUT}\n")


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    print("Loading data...")
    cr_df, vc_df = load_data()

    print("Finding conversion rate increases...")
    cr_increases = find_cr_increases(cr_df)

    print("Finding months with Deal or Coupon rebates...")
    rebate_months = find_rebate_months(vc_df)

    print("Splitting results by rebate presence...")
    no_rebate, with_rebate = split_by_rebate(cr_increases, rebate_months)

    rebate_detail = build_rebate_detail(vc_df)
    brand_summary = build_brand_summary(no_rebate)

    print("Saving formatted output file...")
    save_output(no_rebate, with_rebate, rebate_detail, brand_summary, cr_increases)

    print_summary(cr_increases, no_rebate, with_rebate, brand_summary)


if __name__ == "__main__":
    main()
